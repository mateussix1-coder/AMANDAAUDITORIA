"""
auditoria_engine.py — FreteScan Pro
Parser baseado em pdfplumber (compatível com Python 3.14).

Layout real confirmado via diagnóstico:

ATUA — uma linha por CTE:
  1752 CT 01/04/26 07:14 TR ... TGL1G79 46,900 23.919,00 24.839,65 510,00 ...
  MONEY[0] = peso (46,900)  → ignorar
  MONEY[1] = EmpresaA (23.919,00)
  MONEY[2] = MotoristaA (24.839,65)

GW — uma linha por CTE:
  001752 01/04/2026 ... 46.900,00 23.919,00 0,00 0,00 23.919,00 394,66 1.817,84 0,00 0,00 24.839,88 ...
  MONEY[0] = peso kg (46.900,00)  → ignorar
  MONEY[1] = Valor frete
  MONEY[4] = Frete tab. (EmpresaB)
  MONEY[-3] = Vl Carreteiro Líquido (MotoristaB)
"""

import re
import io
import json
import tempfile
import os

import pdfplumber
import pandas as pd
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime
from pathlib import Path

CENTAVOS = Decimal("0.01")

# Regex para linha de CTE do ATUA:  "1752 CT ..."
RE_ATUA = re.compile(r"^\s*(\d{4,6})\s+CT\b")

# Regex para linha de CTE do GW: "001752 01/04/2026 ..."
RE_GW = re.compile(r"^\s*(\d{4,6})\s+\d{2}/\d{2}/\d{4}\b")

# Captura valores monetários brasileiros, incluindo negativos
MONEY_RE = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2}")

# Peso no ATUA sempre vem como primeiro valor monetário na linha → sempre ignorar índice 0
# (pode ter 2 ou 3 casas decimais dependendo da truncagem da linha)


def parse_money_br(value):
    """Converte moeda brasileira: 23.919,00 → Decimal('23919.00')"""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value.quantize(CENTAVOS, rounding=ROUND_HALF_UP)
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if text in ("", "-", "None", "nan"):
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return Decimal(text).quantize(CENTAVOS, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def normalizar_cte(value):
    text = str(value).strip()
    if not re.fullmatch(r"\d{4,6}", text):
        return None
    numero = int(text)
    if numero < 1000 or numero > 999999:
        return None
    return str(numero)


def _extrair_linhas_pdfplumber(caminho_pdf):
    """Extrai linhas de texto usando pdfplumber (Python 3.14 compatível)."""
    linhas = []
    with pdfplumber.open(str(caminho_pdf)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text() or ""
            for raw in texto.splitlines():
                t = raw.strip()
                if t:
                    linhas.append((page_num, t))
    return linhas


def extrair_atua(caminho_pdf):
    """
    Parser ATUA — layout linha única:
      1752 CT 01/04/26 07:14 TR ... TGL1G79 46,900 23.919,00 24.839,65 510,00 ...

    Valores monetários na linha (MONEY_RE):
      [0] = peso (ex: 46,900)  → ignorar (3 casas decimais)
      [1] = EmpresaA
      [2] = MotoristaA
    """
    registros = {}
    for page_num, linha in _extrair_linhas_pdfplumber(caminho_pdf):
        m = RE_ATUA.match(linha)
        if not m:
            continue
        cte = normalizar_cte(m.group(1))
        if not cte:
            continue

        todos = MONEY_RE.findall(linha)

        # O primeiro valor monetário na linha do ATUA é SEMPRE o peso (ex: 46,900 ou 46,90)
        # EmpresaA = índice 1, MotoristaA = índice 2
        if len(todos) < 3:
            continue

        empresa_a = parse_money_br(todos[1])
        motorista_a = parse_money_br(todos[2])

        if empresa_a is None or motorista_a is None:
            continue

        registros[cte] = {
            "cte": cte,
            "empresa": empresa_a,
            "motorista": motorista_a,
            "pagina": page_num,
        }
    return registros


def extrair_gw(caminho_pdf):
    """
    Parser GW — layout linha única:
      001752 01/04/2026 ... 46.900,00 23.919,00 0,00 0,00 23.919,00 394,66 1.817,84 0,00 0,00 24.839,88 ...

    Valores monetários na linha (MONEY_RE) — índice após filtrar peso:
      [0] = peso kg (46.900,00) → ignorar
      [1] = Valor frete
      [2] = ICMS
      [3] = ICMS%
      [4] = Frete tab. (EmpresaB)
      [5..7] = PIS, COFINS, IR, CSSL
      [-3] = Vl Carreteiro Líquido (MotoristaB)

    Mais robusto: EmpresaB = MONEY[4], MotoristaB = MONEY[-3]
    """
    registros = {}
    for page_num, linha in _extrair_linhas_pdfplumber(caminho_pdf):
        m = RE_GW.match(linha)
        if not m:
            continue
        cte = normalizar_cte(m.group(1))
        if not cte:
            continue

        valores = MONEY_RE.findall(linha)

        # Precisa ter pelo menos 5 valores (peso + frete + icms% + frete_tab + ...)
        if len(valores) < 5:
            continue

        # valores[0] = peso em kg → ignorar
        # valores[4] = Frete tab. = EmpresaB
        # valores[-3] = Vl Carreteiro Líquido = MotoristaB
        empresa_b = parse_money_br(valores[4])
        motorista_b = parse_money_br(valores[-3])

        if empresa_b is None or motorista_b is None:
            continue

        registros[cte] = {
            "cte": cte,
            "empresa": empresa_b,
            "motorista": motorista_b,
            "pagina": page_num,
        }
    return registros


# ---------------------------------------------------------------------------
# Wrappers para o app.py (recebem bytes, escrevem tmp, retornam DataFrame)
# ---------------------------------------------------------------------------

def _bytes_para_tmp(pdf_bytes: bytes) -> str:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_bytes)
        return tmp.name


def parse_atua(pdf_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    tmp = _bytes_para_tmp(pdf_bytes)
    try:
        registros = extrair_atua(tmp)
    finally:
        os.unlink(tmp)

    linhas = [
        {"CTE": r["cte"], "EmpresaA": r["empresa"], "MotoristaA": r["motorista"]}
        for r in registros.values()
    ]
    return pd.DataFrame(linhas), []


def parse_gw(pdf_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    tmp = _bytes_para_tmp(pdf_bytes)
    try:
        registros = extrair_gw(tmp)
    finally:
        os.unlink(tmp)

    linhas = [
        {"CTE": r["cte"], "EmpresaB": r["empresa"], "MotoristaB": r["motorista"]}
        for r in registros.values()
    ]
    return pd.DataFrame(linhas), []


# ---------------------------------------------------------------------------
# Lógica de auditoria
# ---------------------------------------------------------------------------

def comparar(df_a: pd.DataFrame, df_b: pd.DataFrame, tolerancia: Decimal) -> pd.DataFrame:
    set_a = set(df_a["CTE"]) if not df_a.empty else set()
    set_b = set(df_b["CTE"]) if not df_b.empty else set()
    dict_a = df_a.set_index("CTE").to_dict("index") if not df_a.empty else {}
    dict_b = df_b.set_index("CTE").to_dict("index") if not df_b.empty else {}

    resultados = []
    for cte in sorted(set_a | set_b, key=lambda x: int(x)):
        row_a = dict_a.get(cte, {})
        row_b = dict_b.get(cte, {})

        ea = row_a.get("EmpresaA")
        ma = row_a.get("MotoristaA")
        eb = row_b.get("EmpresaB")
        mb = row_b.get("MotoristaB")

        em_a = ea is not None and ma is not None
        em_b = eb is not None and mb is not None

        if em_a and em_b:
            de = (ea - eb).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            dm = (ma - mb).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            maior = max(abs(de), abs(dm))
            if de == 0 and dm == 0:
                status = "OK"
            elif abs(de) <= tolerancia and abs(dm) <= tolerancia:
                status = "OK por arredondamento"
            else:
                status = "Divergente"
        elif em_a:
            status = "Faltante no B"
            de = dm = maior = None
        else:
            status = "Faltante no A"
            de = dm = maior = None

        resultados.append({
            "CTE": cte,
            "Status": status,
            "Empresa A": float(ea) if ea is not None else None,
            "Motorista A": float(ma) if ma is not None else None,
            "Empresa B": float(eb) if eb is not None else None,
            "Motorista B": float(mb) if mb is not None else None,
            "Dif. Empresa": float(de) if de is not None else None,
            "Dif. Motorista": float(dm) if dm is not None else None,
            "Maior Diferença": float(maior) if maior is not None else None,
        })

    return pd.DataFrame(resultados)


def gerar_resumo(df: pd.DataFrame) -> dict:
    total = len(df)
    ok = len(df[df["Status"] == "OK"])
    ok_r = len(df[df["Status"] == "OK por arredondamento"])
    div = len(df[df["Status"] == "Divergente"])
    fa = len(df[df["Status"] == "Faltante no A"])
    fb = len(df[df["Status"] == "Faltante no B"])
    dfx = df[df["Status"] == "Divergente"]
    return {
        "total": total,
        "ok": ok,
        "ok_arredondamento": ok_r,
        "divergentes": div,
        "faltantes_a": fa,
        "faltantes_b": fb,
        "dif_total_empresa": round(dfx["Dif. Empresa"].sum(), 2) if not dfx.empty else 0.0,
        "dif_total_motorista": round(dfx["Dif. Motorista"].sum(), 2) if not dfx.empty else 0.0,
        "impacto_absoluto": round(dfx["Maior Diferença"].sum(), 2) if not dfx.empty else 0.0,
    }


# ---------------------------------------------------------------------------
# Exportação
# ---------------------------------------------------------------------------

def exportar_csv(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.to_csv(buf, sep=";", index=False, encoding="utf-8-sig")
    return buf.getvalue().encode("utf-8-sig")


def exportar_excel(df, resumo, nome_a, nome_b, tolerancia) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    CORES = {
        "OK": "C6EFCE", "OK por arredondamento": "FFEB9C",
        "Divergente": "FFC7CE", "Faltante no A": "DDEBF7", "Faltante no B": "FCE4D6",
    }
    buf = io.BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.append(["FreteScan Pro — Resumo de Auditoria"])
    ws1.append([f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws1.append([f"Arquivo A: {nome_a}  |  Arquivo B: {nome_b}"])
    ws1.append([f"Tolerância: R$ {tolerancia:.2f}"])
    ws1.append([])
    ws1.append(["Métrica", "Valor"])
    for k, v in resumo.items():
        ws1.append([k, v])

    ws2 = wb.create_sheet("Auditoria")
    ws2.append(list(df.columns))
    ws2.freeze_panes = "A2"
    for r, row in enumerate(df.itertuples(index=False), start=2):
        ws2.append([x if pd.notna(x) else "" for x in row])
        fill = PatternFill(
            start_color=CORES.get(row.Status, "FFFFFF"),
            end_color=CORES.get(row.Status, "FFFFFF"),
            fill_type="solid",
        )
        for c in range(1, len(df.columns) + 1):
            ws2.cell(row=r, column=c).fill = fill
    for col in ws2.columns:
        w = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    wb.save(buf)
    return buf.getvalue()


def exportar_pdf(df, resumo, nome_a, nome_b, tolerancia) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph("FreteScan Pro — Relatório de Auditoria", styles["Title"]))
    elems.append(Spacer(1, 0.3 * cm))
    elems.append(Paragraph(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles["Normal"]))
    elems.append(Paragraph(f"Arquivo A: {nome_a}  |  Arquivo B: {nome_b}", styles["Normal"]))
    elems.append(Paragraph(f"Tolerância: R$ {tolerancia:.2f}  —  Diferença = A − B", styles["Normal"]))
    elems.append(Spacer(1, 0.5 * cm))

    elems.append(Paragraph("Resumo Geral", styles["Heading2"]))
    res_data = [["Métrica", "Valor"]] + [[k, str(v)] for k, v in resumo.items()]
    t = Table(res_data, colWidths=[8 * cm, 4 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 0.5 * cm))

    df_div = df[df["Status"].isin(["Divergente", "Faltante no A", "Faltante no B"])].head(50)
    if not df_div.empty:
        elems.append(Paragraph("Divergências e Faltantes (até 50)", styles["Heading2"]))
        cols = ["CTE", "Status", "Empresa A", "Empresa B", "Dif. Empresa",
                "Motorista A", "Motorista B", "Dif. Motorista"]
        data = [cols] + [
            ["" if pd.isna(row[c]) else str(row[c]) for c in cols]
            for _, row in df_div.iterrows()
        ]
        cw = [2 * cm, 4 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm]
        t2 = Table(data, colWidths=cw, repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        elems.append(t2)

    doc.build(elems)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Histórico
# ---------------------------------------------------------------------------

HIST_PATH = Path("historico_auditoria.json")


def salvar_historico(nome_a, nome_b, tolerancia, resumo):
    hist = []
    if HIST_PATH.exists():
        try:
            hist = json.loads(HIST_PATH.read_text(encoding="utf-8"))
        except Exception:
            hist = []
    hist.insert(0, {
        "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "arquivo_a": nome_a,
        "arquivo_b": nome_b,
        "tolerancia": float(tolerancia),
        **resumo,
    })
    HIST_PATH.write_text(json.dumps(hist[:100], ensure_ascii=False, indent=2), encoding="utf-8")


def carregar_historico() -> list:
    if not HIST_PATH.exists():
        return []
    try:
        return json.loads(HIST_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []
